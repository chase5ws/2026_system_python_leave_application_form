import openpyxl
from openpyxl.drawing.image import Image
import tkinter as tk
from tkinter import ttk, Text, messagebox, filedialog
from datetime import datetime, date, timedelta
import os
import shutil
from tkcalendar import DateEntry
from PIL import Image as PILImage

class LeaveApplicationSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("員工請假單生成程序 By ChaseTseng")
        self.signature_path = None
        self.signature_cell = "I19"

        self.signature_type = tk.StringVar(value="本人簽名")
        
        try:
            self.root.iconbitmap("my_icon.ico")
        except Exception:
            pass
        
        self.EXCEL_PATH, self.COPY_EXCEL_PATH = self.load_config()
        self.employee_data, self.employee_names = self.load_employee_data()
        
        self.selected_leave_type = tk.StringVar(value="")
        self.current_employee = {"部門": "", "姓名": "", "工號": ""}
        
        self.build_ui()

    def load_config(self):
        config_path = "Leave_Application_Form_config.txt"
        excel_path = ""
        copy_excel_path = ""
        
        if not os.path.exists(config_path):
            with open(config_path, 'w', encoding='utf-8') as f:
                f.write('Leave_Application_Form_EXCEL_PATH = "火影員工名單.xlsx"\n')
                f.write('Leave_Application_Form_COPY_EXCEL_PATH = "火影請假單 Leave App Form.xlsx"\n')
            excel_path = "火影員工名單.xlsx"
            copy_excel_path = "火影請假單 Leave App Form.xlsx"
        else:
            with open(config_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                for line in lines:
                    line = line.strip()
                    if line.startswith("Leave_Application_Form_EXCEL_PATH"):
                        excel_path = line.split("=")[1].strip().strip('"')
                    elif line.startswith("Leave_Application_Form_COPY_EXCEL_PATH"):
                        copy_excel_path = line.split("=")[1].strip().strip('"')
        
        return excel_path, copy_excel_path

    def load_employee_data(self):
        employee_dict = {}
        employee_names = []
        
        try:
            wb = openpyxl.load_workbook(self.EXCEL_PATH)
            ws = wb.active
            
            header_row = [cell.value for cell in ws[1]]
            required_headers = ["部門", "姓名", "工號"]
            for header in required_headers:
                if header not in header_row:
                    raise ValueError(f"Excel缺少標題：{header}")
            
            dept_col = header_row.index("部門")
            name_col = header_row.index("姓名")
            id_col = header_row.index("工號")
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[name_col] is None:
                    continue
                
                dept = str(row[dept_col]) if row[dept_col] else ""
                name = str(row[name_col])
                emp_id = str(row[id_col]) if row[id_col] else ""
                
                employee_dict[name] = {"部門": dept, "工號": emp_id}
                employee_names.append(name)
            
            wb.close()
            
        except FileNotFoundError:
            raise FileNotFoundError(f"找不到 {self.EXCEL_PATH}")
        except Exception as e:
            raise Exception(f"讀取員工清單失敗：{str(e)}")
        
        return employee_dict, employee_names

    def on_name_selected(self, event):
        selected_name = self.name_combobox.get()
        if selected_name in self.employee_data:
            self.current_employee = {
                "姓名": selected_name,
                "部門": self.employee_data[selected_name]["部門"],
                "工號": self.employee_data[selected_name]["工號"]
            }
            self.dept_label.config(text=f"部門：{self.current_employee['部門']}")
            self.emp_id_label.config(text=f"工號：{self.current_employee['工號']}")
        else:
            self.dept_label.config(text="部門：")
            self.emp_id_label.config(text="工號：")
            self.current_employee = {"部門": "", "姓名": "", "工號": ""}

    def on_leave_selected(self, leave_type):
        try:
            self.selected_leave_type.set(leave_type)
            for rb in self.leave_buttons:
                if rb["text"] != leave_type and rb.instate(["selected"]):
                    rb.deselect()
        except:
            pass

    def calculate_working_days(self, start_date, end_date):
        working_days = 0
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5:
                working_days += 1
            current_date += timedelta(days=1)
        return working_days

    # ===================== 這裡重寫：8小時以下只顯示小時 =====================
    def calculate_total_hours_auto(self, *args):
        try:
            start_date = self.start_calendar.get_date()
            end_date = self.end_calendar.get_date()
            
            s_h = int(self.start_hour_combobox.get())
            s_m = int(self.start_minute_combobox.get())
            e_h = int(self.end_hour_combobox.get())
            e_m = int(self.end_minute_combobox.get())
            
            if end_date < start_date:
                self.end_calendar.set_date(start_date)
                end_date = start_date

            total_hours = 0.0
            current_date = start_date
            
            while current_date <= end_date:
                # 跳过周六、周日
                if current_date.weekday() >= 5:
                    current_date += timedelta(days=1)
                    continue
                
                # 定义当天的工作时间边界
                work_start = timedelta(hours=8, minutes=30)
                work_end = timedelta(hours=17, minutes=30)
                lunch_start = timedelta(hours=12, minutes=30)
                lunch_end = timedelta(hours=13, minutes=30)
                
                # 获取当天的请假起止时间
                if current_date == start_date:
                    t_start = timedelta(hours=s_h, minutes=s_m)
                else:
                    t_start = work_start
                    
                if current_date == end_date:
                    t_end = timedelta(hours=e_h, minutes=e_m)
                else:
                    t_end = work_end
                
                # 计算上午时长（8:30 - 12:30）
                morning_start = max(t_start, work_start)
                morning_end = min(t_end, lunch_start)
                morning_hours = (morning_end - morning_start).total_seconds() / 3600
                morning_hours = max(morning_hours, 0)
                
                # 计算下午时长（13:30 - 17:30）
                afternoon_start = max(t_start, lunch_end)
                afternoon_end = min(t_end, work_end)
                afternoon_hours = (afternoon_end - afternoon_start).total_seconds() / 3600
                afternoon_hours = max(afternoon_hours, 0)
                
                # 累加当天有效时长
                total_hours += morning_hours + afternoon_hours
                current_date += timedelta(days=1)

            total_hours = round(total_hours, 1)
            
            if total_hours <= 8:
                self.total_label.config(text=f"{total_hours} 小時")
            else:
                days = int(total_hours // 8)
                rem = round(total_hours % 8, 1)
                if rem == 0:
                    self.total_label.config(text=f"{days} 天（{total_hours:.0f} 小時）")
                else:
                    self.total_label.config(text=f"{days} 天 {rem} 小時（{total_hours:.1f} 小時）")

        except Exception as e:
            self.total_label.config(text="0 小時")
            print(f"计算错误: {e}")




    def convert_image_to_png(self):
        input_path = filedialog.askopenfilename(
            title="選擇要轉換的圖片檔案",
            filetypes=[("所有圖片檔案", "*.jpg *.jpeg *.png *.bmp *.gif *.tiff *.webp"),
                       ("PNG檔案", "*.png"), ("JPG檔案", "*.jpg *.jpeg")]
        )
        if not input_path:
            return
        if not os.path.exists(input_path):
            messagebox.showerror("錯誤", "選取的檔案不存在！")
            return
        try:
            with PILImage.open(input_path) as img:
                if getattr(img, "is_animated", False):
                    img = img.convert("RGBA")
                fname = os.path.splitext(os.path.basename(input_path))[0]
                save_path = filedialog.asksaveasfilename(
                    title="儲存PNG檔案", defaultextension=".png",
                    initialfile=fname, filetypes=[("PNG檔案", "*.png")]
                )
                if save_path:
                    img.save(save_path, format="PNG")
                    messagebox.showinfo("成功", f"圖片已轉換：\n{save_path}")
        except Exception as e:
            messagebox.showerror("轉換失敗", f"錯誤：{str(e)}")

    def upload_signature(self):
        file_path = filedialog.askopenfilename(
            title="選擇簽名圖片",
            filetypes=[("所有圖片檔案", "*.jpg *.jpeg *.png *.bmp *.gif *.tiff *.webp"),
                       ("PNG檔案", "*.png"), ("JPG檔案", "*.jpg *.jpeg")]
        )
        if file_path:
            self.signature_path = file_path
            self.signature_label.config(text=f"已載入：{os.path.basename(file_path)} → 自動存入 I19")
        else:
            self.signature_path = None
            self.signature_label.config(text="未選擇簽名")

    def write_to_cell(self, cell, value):
        for mr in cell.parent.merged_cells.ranges:
            if cell.coordinate in mr:
                cell.parent[mr.start_cell.coordinate] = value
                return
        cell.value = value

    def clear_bottom_formatting(self, ws):
        for col in range(1, 11):
            for row in range(37, ws.max_row + 1):
                c = ws.cell(row=row, column=col)
                c.border = openpyxl.styles.Border()
                c.fill = openpyxl.styles.PatternFill(fill_type=None)
                c.font = openpyxl.styles.Font()
                c.alignment = openpyxl.styles.Alignment()

    def generate_leave_form(self):
        if not self.current_employee["姓名"]:
            messagebox.showwarning("警告", "請先選擇員工！")
            return
        if not self.selected_leave_type.get():
            messagebox.showwarning("警告", "請選擇假別！")
            return
        if not os.path.exists(self.COPY_EXCEL_PATH):
            messagebox.showerror("錯誤", f"找不到模板 {self.COPY_EXCEL_PATH}")
            return

        self.calculate_total_hours_auto()

        try:
            emp_name = self.current_employee["姓名"]
            today_str = datetime.now().strftime("%Y%m%d")
            base = f"請假單 Leave App Form_{emp_name}_{today_str}"

            save_dir = filedialog.askdirectory(title="選擇儲存資料夾")
            if not save_dir:
                messagebox.showinfo("提示", "已取消")
                return

            for suffix, label_text in [("人事部留存", "人事部留存"), ("申請人留存", "申請人留存")]:
                fn = f"{base}_{suffix}.xlsx"
                fp = os.path.join(save_dir, fn)
                shutil.copy2(self.COPY_EXCEL_PATH, fp)

                wb = openpyxl.load_workbook(fp)
                ws = wb.active

                name = self.current_employee["姓名"]
                emp_id = self.current_employee["工號"]
                dept = self.current_employee["部門"]
                leave_type = self.selected_leave_type.get()
                desc = self.desc_text.get("1.0", tk.END).strip()
                
                s_date = self.start_calendar.get_date().strftime("%Y/%m/%d")
                s_h = self.start_hour_combobox.get()
                s_m = self.start_minute_combobox.get()
                e_date = self.end_calendar.get_date().strftime("%Y/%m/%d")
                e_h = self.end_hour_combobox.get()
                e_m = self.end_minute_combobox.get()
                
                start_full = f"{s_date} {s_h}:{s_m}"
                end_full = f"{e_date} {e_h}:{e_m}"
                period = f"{start_full} 至 {end_full}"
                
                total = self.total_label.cget("text")
                app_date = datetime.now().strftime("%Y/%m/%d")

                self.write_to_cell(ws["E5"], "申請日期 DATE:")
                self.write_to_cell(ws["G5"], "申請日期")
                self.write_to_cell(ws["I5"], app_date)
                self.write_to_cell(ws["B6"], name)
                self.write_to_cell(ws["E6"], emp_id)
                self.write_to_cell(ws["I6"], dept)
                self.write_to_cell(ws["B8"], leave_type)
                self.write_to_cell(ws["B12"], desc)
                
                if self.signature_type.get() == "本人簽名":
                    self.write_to_cell(ws["G12"], "本人簽名 Applicant's Signature")
                else:
                    self.write_to_cell(ws["G12"], "代理人簽名 Signature of Acting Person")
                    
                self.write_to_cell(ws["B24"], period)
                self.write_to_cell(ws["I24"], total)
                self.write_to_cell(ws["A36"], label_text)

                if self.signature_path and os.path.exists(self.signature_path):
                    try:
                        img = Image(self.signature_path)
                        img.width = 120
                        img.height = 60
                        ws.add_image(img, self.signature_cell)
                    except Exception:
                        messagebox.showwarning("提示", "插入簽名失敗")

                self.clear_bottom_formatting(ws)
                wb.save(fp)
                wb.close()

            messagebox.showinfo("完成", f"已產生 2 個檔案：\n{base}_人事部留存.xlsx\n{base}_申請人留存.xlsx")

        except PermissionError:
            messagebox.showerror("錯誤", "檔案被開啟，請關閉後再試")
        except Exception as e:
            messagebox.showerror("錯誤", f"失敗：{str(e)}")

    def build_ui(self):
        pad = {"padx": 10, "pady": 5}

        frame1 = ttk.Frame(self.root)
        frame1.pack(fill="x", **pad)
        ttk.Label(frame1, text="員工：").grid(row=0, column=0, **pad)
        self.name_combobox = ttk.Combobox(frame1, values=self.employee_names, state="readonly")
        self.name_combobox.grid(row=0, column=1, **pad)
        self.name_combobox.bind("<<ComboboxSelected>>", self.on_name_selected)
        self.dept_label = ttk.Label(frame1, text="部門：")
        self.dept_label.grid(row=0, column=2, **pad)
        self.emp_id_label = ttk.Label(frame1, text="工號：")
        self.emp_id_label.grid(row=0, column=3, **pad)

        frame1_1 = ttk.Frame(self.root)
        frame1_1.pack(fill="x", **pad)
        ttk.Label(frame1_1, text="簽名：").grid(row=0, column=0, **pad)
        rb1 = ttk.Radiobutton(frame1_1, text="本人簽名Applicant's Signature", 
                              variable=self.signature_type, value="本人簽名")
        rb1.grid(row=0, column=1, sticky="w", **pad)
        rb1.invoke()
        rb2 = ttk.Radiobutton(frame1_1, text="代理人簽名Signature of Acting Person", 
                              variable=self.signature_type, value="代理人簽名")
        rb2.grid(row=0, column=2, sticky="w", **pad)

        frame2 = ttk.Frame(self.root)
        frame2.pack(fill="x", **pad)
        ttk.Label(frame2, text="假別：").grid(row=0, column=0, **pad, rowspan=4)
        leave_types = [
            "年假 Vacation Leave", "婚假 Vacation Leave", "補休 Compensatory Leave",
            "事假 Personal Leave", "產假 Maternity Leave", "喪假 Bereavement Leave",
            "普通傷病假 Sick Leave", "陪產假 Paternity Leave", "公假 Official Leave",
            "生理假 Menstruation Leave", "家庭照顧假 Family Care Leave", "其他 Other Leave"
        ]
        self.leave_buttons = []
        c, r = 1, 0
        for lt in leave_types:
            rb = ttk.Radiobutton(frame2, text=lt, variable=self.selected_leave_type,
                                 value=lt, command=lambda x=lt: self.on_leave_selected(x))
            rb.grid(row=r, column=c, sticky="w", **pad)
            self.leave_buttons.append(rb)
            c += 1
            if c > 3:
                c, r = 1, r+1

        frame3 = ttk.Frame(self.root)
        frame3.pack(fill="both", expand=True, **pad)
        ttk.Label(frame3, text="說明：").grid(row=0, column=0, sticky="n", **pad)
        self.desc_text = Text(frame3, width=50, height=5)
        self.desc_text.grid(row=0, column=1, sticky="nsew", **pad)

        sig_frame = ttk.Frame(frame3)
        sig_frame.grid(row=0, column=2, sticky="n", **pad)
        ttk.Button(sig_frame, text="圖片轉PNG格式", command=self.convert_image_to_png).pack(fill="x", **pad)
        ttk.Button(sig_frame, text="上傳簽名檔", command=self.upload_signature).pack(fill="x", **pad)
        self.signature_label = ttk.Label(sig_frame, text="未選擇簽名")
        self.signature_label.pack(fill="x", **pad)

        frame4 = ttk.Frame(self.root)
        frame4.pack(fill="x", **pad)
        
        ttk.Label(frame4, text="開始：").grid(row=0, column=0, **pad)
        self.start_calendar = DateEntry(frame4, width=15, date_pattern="yyyy/mm/dd")
        self.start_calendar.set_date(date.today())
        self.start_calendar.grid(row=0, column=1, **pad)
        
        self.start_hour_combobox = ttk.Combobox(frame4, values=[f"{h:02d}" for h in range(24)], width=3, state="readonly")
        self.start_hour_combobox.set("08")
        self.start_hour_combobox.grid(row=0, column=2, padx=(0,2))
        ttk.Label(frame4, text=":").grid(row=0, column=3, padx=2)
        self.start_minute_combobox = ttk.Combobox(frame4, values=["00","15","30","45"], width=3, state="readonly")
        self.start_minute_combobox.set("30")
        self.start_minute_combobox.grid(row=0, column=4, padx=(2,10))

        ttk.Label(frame4, text="結束：").grid(row=0, column=5, **pad)
        self.end_calendar = DateEntry(frame4, width=15, date_pattern="yyyy/mm/dd")
        self.end_calendar.set_date(date.today())
        self.end_calendar.grid(row=0, column=6, **pad)
        
        self.end_hour_combobox = ttk.Combobox(frame4, values=[f"{h:02d}" for h in range(24)], width=3, state="readonly")
        self.end_hour_combobox.set("17")
        self.end_hour_combobox.grid(row=0, column=7, padx=(0,2))
        ttk.Label(frame4, text=":").grid(row=0, column=8, padx=2)
        self.end_minute_combobox = ttk.Combobox(frame4, values=["00","15","30","45"], width=3, state="readonly")
        self.end_minute_combobox.set("30")
        self.end_minute_combobox.grid(row=0, column=9, padx=(2,10))

        self.start_calendar.bind("<<DateEntrySelected>>", self.calculate_total_hours_auto)
        self.end_calendar.bind("<<DateEntrySelected>>", self.calculate_total_hours_auto)
        self.start_hour_combobox.bind("<<ComboboxSelected>>", self.calculate_total_hours_auto)
        self.start_minute_combobox.bind("<<ComboboxSelected>>", self.calculate_total_hours_auto)
        self.end_hour_combobox.bind("<<ComboboxSelected>>", self.calculate_total_hours_auto)
        self.end_minute_combobox.bind("<<ComboboxSelected>>", self.calculate_total_hours_auto)

        frame5 = ttk.Frame(self.root)
        frame5.pack(fill="x", **pad)
        ttk.Label(frame5, text="💡 提醒：需要先選擇開始，再選擇結束時間才會正確", 
                  font=("Microsoft JhengHei",10,"bold"), foreground="black").grid(
            row=0, column=0, columnspan=2, sticky=tk.W, pady=(0,5))
        
        ttk.Label(frame5, text="請假合計：", font=("Microsoft JhengHei",10)).grid(row=1, column=0, sticky=tk.W, **pad)
        self.total_label = ttk.Label(frame5, text="8 小時", font=("Microsoft JhengHei",10))
        self.total_label.grid(row=1, column=1, sticky=tk.W, **pad)

        frame6 = ttk.Frame(self.root)
        frame6.pack(fill="x", **pad)
        ttk.Button(frame6, text="生成請假單", command=self.generate_leave_form).grid(row=0, column=0, **pad)

if __name__ == "__main__":
    root = tk.Tk()
    app = LeaveApplicationSystem(root)
    app.calculate_total_hours_auto()
    root.mainloop()